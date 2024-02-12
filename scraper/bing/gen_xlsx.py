#!/usr/bin/env python
#  -*- coding:utf-8 -*-

import openpyxl
import os
import time
import bing_dict

def convert_txt_to_xlsx(txt_file, xlsx_file):
    # 打开txt文件并读取单词列表
    with open(txt_file, 'r', encoding='utf-8') as txt:
        word_list = txt.read().splitlines()

    # 如果有进度文件，从上次进度继续，否则从头开始
    progress_file_name = f"{xlsx_file}.process"
    if os.path.exists(progress_file_name):
        workbook = openpyxl.load_workbook(xlsx_file)
        with open(progress_file_name, 'r') as progress_file:
            progress = int(progress_file.read())
    else:
        workbook = openpyxl.Workbook()
        progress = 0

    print(f"Last process: {progress}")

    # 创建词典提取器
    de = bing_dict.DictExtractor()

    # 设置标题行
    worksheet = workbook.active
    worksheet['A1'] = '单词'
    worksheet['B1'] = '音标'
    worksheet['C1'] = '释义'
    worksheet['D1'] = '评分'
    worksheet['E1'] = '例句'

    # 创建一个 Alignment 对象，设置对齐方式
    alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=False)
                    
    # 将单词列表写入Excel表格
    for i, word in enumerate(word_list):
        seq = i + 1

        # 从上次进度开始
        if seq <= progress:
            continue

        print(f'Excel: {seq}/{len(word_list)} {word}')

        word_info = de.extract_word(word)
        worksheet.cell(row=seq+1, column=1, value=word_info.word_text()).alignment = alignment
        worksheet.cell(row=seq+1, column=2, value=word_info.pronunciation_text()).alignment = alignment
        worksheet.cell(row=seq+1, column=3, value=word_info.definition_text()).alignment = alignment
        worksheet.cell(row=seq+1, column=4, value=1).alignment = alignment
        worksheet.cell(row=seq+1, column=5, value=word_info.example_text()).alignment = alignment

        # 分批次保存一下，在有大量数据情况下，有时异常中断时，就不用完全从头开始
        if seq % 10 == 0:
            # 保存数据文件
            workbook.save(xlsx_file)
            # 记录当前进度
            with open(progress_file_name, 'w') as progress_file:
                progress_file.write(str(seq))

        # 分批次休息一下，尽量避免连续大量请求
        if seq % 100 == 0:
            time.sleep(5)

    # 完整结束后，保存文件，并清除进度文件
    workbook.save(xlsx_file)
    os.remove(progress_file_name)

if __name__ == '__main__':
    txt_file = 'input.txt'  # 请将文件名替换为您的txt文件名
    xlsx_file = 'output.xlsx'  # 请将文件名替换为您的xlsx文件名
    retry_max = 10 # 遇到异常最大重试次数

    # 执行功能，遇到异常自动重试，直到达到最大次数
    for retry_count in range(retry_max):
        try:
            result = convert_txt_to_xlsx(txt_file, xlsx_file)
            print(f"[Info] 完成")
            break
        except Exception as e:
            print(f"[Warning] 尝试 {retry_count + 1} 出现异常: {str(e)}")
 