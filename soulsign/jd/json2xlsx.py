#!/usr/bin/env python
#  -*- coding:utf-8 -*-

import io
import json
import sys
import time
import urllib.parse
import urllib.request
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')

def readOrdersFromJsonFile(jsonPath):
    orders = []
    with open(jsonPath, encoding='utf-8') as f:
        orders = json.load(f)

    print('Read orders from json file[%s] done' % jsonPath)
    return orders

def writeIntoExcelFile(excelPath, orders):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '京东我的订单'
    ws.cell(1, 1, '订单ID')
    ws.cell(1, 2, '订单时间')
    ws.cell(1, 3, '订单金额')
    ws.cell(1, 4, '商品名称')
    ws.cell(1, 5, '商品价格')
    ws.cell(1, 6, '商品数量')

    line = 1
    for order in orders:
        line += 1
        ws.cell(line, 1, order['orderId'])
        ws.cell(line, 2, order['dealTime'])
        ws.cell(line, 3, order['costMoney'])
        ws.cell(line, 4, order['itemName'])
        ws.cell(line, 5, order['itemPrice'])
        ws.cell(line, 6, order['itemAmount'])
        print(order)

    wb.save(excelPath)
    print('Write into excel file[%s] done' % excelPath)

# Current time, ingore seconds
now = int(time.time())
now -= now % 60

jsonPath = 'jd_my_orders.json'
orders = readOrdersFromJsonFile(jsonPath)

statTime = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(now))
excelPath = 'jd_my_orders_' + statTime + '.xlsx'
writeIntoExcelFile(excelPath, orders)
