#!/usr/bin/env python
#  -*- coding:utf-8 -*-

import configparser
import datetime
import io
import openpyxl
import os
import pymysql
import re
import requests
import sys
import time

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')

# Current time, ingore seconds
nowSecond = int(time.time())
nowSecond -= nowSecond % 60
nowDatetime = datetime.datetime.utcfromtimestamp(nowSecond).strftime("%Y-%m-%d %H:%M:%S")
print('Now second: %d datetime: %s' % (nowSecond, nowDatetime))

# Current Directory
curDir = os.path.dirname(os.path.abspath(__file__))
print('Current path: %s' % (curDir))

major = 'worldometers'
minor = 'country gdp'

url = 'https://www.worldometers.info/gdp/'
headers = {
    'Referer': 'https://www.worldometers.info/',
    'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 11_4_1 like Mac OS X) AppleWebKit/600.1.4 (KHTML, like Gecko) Mobile/12A365 QQMusic/8.6.0 Mskin/white Mcolor/31c27cff Bcolor/00000000 skinid[0] NetType/WIFI WebView/UIWebView Released[1] zh-CN DeviceModel/iPhone7,2',
}

records = []

def fetchDataFromPage(records):
    rsp = requests.get(url, headers = headers)
    text = rsp.text

    blockPattern = r't20-line">.*?</div>'
    blockMatches = re.finditer(blockPattern, text)
    for blockSeq, blockMatch in enumerate(blockMatches, start=1):
        blockText = blockMatch.group(0)
   
        unitPattern = r'">([^>]+)</a></span> <span class="t20-number-gdp">\$([^>]+)</span>'
        unitMatches = re.findall(unitPattern, blockText)
        
        country = unitMatches[0][0]
        gdp = unitMatches[0][1]

        record = {}
        record['seq'] = blockSeq
        record['country'] = country
        record['gdp'] = int(gdp.replace(',', ''))
        record['datetime'] = nowDatetime
        records.append(record)

fetchDataFromPage(records)

def writeIntoFile(fileName, records):
    xlsDir = 'data'
    xlsName = '%s.xlsx' % fileName
    xlsPath = os.path.abspath(os.path.join(curDir, xlsDir, xlsName))
    print('Xls file path: %s' % (xlsPath))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '统计'
    ws.cell(1, 1, '序号')
    ws.cell(1, 2, '国家/地区')
    ws.cell(1, 3, 'GDP')
    ws.cell(1, 4, '时间')

    line = 1
    for record in records:
        line += 1
        ws.cell(line, 1, record['seq'])
        ws.cell(line, 2, record['country'])
        ws.cell(line, 3, record['gdp'])
        ws.cell(line, 4, record['datetime'])

    wb.save(xlsPath)
    print('Write into file xls(%s) done' % xlsPath)

statTime = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(nowSecond))
fileName = 'country_gdp_%s' % statTime
writeIntoFile(fileName, records)

def writeIntoDB(tableName, records):
    confPath = os.path.abspath(os.path.join(curDir, '../conf/db.conf'))
    print('DB conf path: %s' % (confPath))

    confDB = configparser.ConfigParser()
    confDB.read(confPath)

    mysqlHost        = confDB.get('mysql', 'host')
    mysqlPort        = confDB.get('mysql', 'port')
    mysqlUser        = confDB.get('mysql', 'user')
    mysqlPassword    = confDB.get('mysql', 'password')
    mysqlDB          = confDB.get('mysql', 'db')
    mysqlCharset     = confDB.get('mysql', 'charset')
    print('DB mysql info: %s %s %s %s %s %s' % (mysqlHost, mysqlPort, mysqlUser, mysqlPassword, mysqlDB, mysqlCharset))

    db = pymysql.connect(host=mysqlHost, port=int(mysqlPort), user=mysqlUser, passwd=mysqlPassword, db=mysqlDB, charset=mysqlCharset)
    cursor = db.cursor()

    for record in records:
        if record['seq'] > 200:
            break

        sql = "INSERT INTO {tableName} VALUES(0, '{factor}', '{major}', '{minor}', '{time}', '{unit}', '{value}')"
        sql = sql.format(tableName = tableName, factor=record['datetime'], major=major, minor=minor, time=nowSecond, unit=record['country'], value=record['gdp'])
        print(sql)
        try:
            cursor.execute(sql)
            db.commit()
        except Exception as e:
            db.rollback()
            print('Insert record(%d) error(%s)!' % record['seq'], str(e))

    db.close()
    print('Write into db table(%s) done' % tableName)

writeIntoDB('tab_generic_trend', records)
